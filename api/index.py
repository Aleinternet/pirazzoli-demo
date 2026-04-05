# C:\ale\zz_ale_cosas\otras_personas\Pirazzoli\pirazzoli-demo\api\index.py

import os
import io
import re
import base64
import json
import requests
from datetime import datetime, date
from flask import Flask, jsonify, request
from openpyxl import load_workbook
from pypdf import PdfReader
import unicodedata
import time

app = Flask(__name__)

TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
SHAREPOINT_HOSTNAME = os.getenv("SHAREPOINT_HOSTNAME")
SHAREPOINT_SITE_PATH = os.getenv("SHAREPOINT_SITE_PATH")
TARGET_DRIVE_NAME = os.getenv("TARGET_DRIVE_NAME", "Documentos")
TARGET_FILE_NAME = os.getenv("TARGET_FILE_NAME", "piloto_datos_minerva.xlsx")
CMF_API_KEY = os.getenv("CMF_API_KEY")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")



API_CACHE = {
    "payload": None,
    "last_build_ts": 0,
    "last_modified": None,
}
CACHE_SECONDS = 14400

PDF_DATE_CACHE = {}
CMF_RATE_CACHE = {}

def build_payload(force_refresh=False):
    now = time.time()

    if (
        not force_refresh
        and API_CACHE["payload"] is not None
        and (now - API_CACHE["last_build_ts"]) < CACHE_SECONDS
    ):
        return API_CACHE["payload"]

    token = get_token()
    excel_files = fetch_excel_files(token)

    files = []
    last_modified_values = []

    for item in excel_files:
        parsed = parse_workbook_to_payload(
            item["content"],
            token,
            file_name=item["file_name"],
            run_audit=force_refresh
        )
        files.extend(parsed)

        if item.get("last_modified"):
            last_modified_values.append(item["last_modified"])

    payload = {
        "ok": True,
        "source": "sharepoint",
        "lastModified": max(last_modified_values) if last_modified_values else None,
        "files": files
    }

    API_CACHE["payload"] = payload
    API_CACHE["last_build_ts"] = now
    API_CACHE["last_modified"] = payload["lastModified"]
    return payload

def normalize_text(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def detect_transport_key(sheet_name: str):
    parsed = detect_sheet_pattern(sheet_name)
    return parsed["transportKey"] if parsed else None


def detect_operation_type(sheet_name: str):
    parsed = detect_sheet_pattern(sheet_name)
    return parsed["operationType"] if parsed else None


def parse_excel_file_name(file_name: str):
    """
    Espera nombres tipo:
    Comex_EMPRESA_AÑO_NUMERO_descripcion.xlsx

    Ejemplos válidos:
    - Comex_Minerva_2026_01_importaciones.xlsx
    - Comex_Minerva_2026_02.xlsx
    - Comex_CocaCola_2025_03_exportaciones_mayo.xlsm

    Devuelve metadata paramétrica sin depender de la descripción.
    """
    original = file_name or ""
    base = original.rsplit(".", 1)[0].strip()

    parts = [p.strip() for p in re.split(r"[_\-]+", base) if p.strip()]
    result = {
        "file_name": original,
        "base_name": base,
        "prefix": "",
        "company_name": "Empresa",
        "year": "",
        "number": "",
        "description": "",
    }

    if not parts:
        return result

    # Debe comenzar con Comex, pero si no, igual intentamos rescatar info.
    if parts:
        result["prefix"] = parts[0]

    # Formato esperado: Comex, EMPRESA, AÑO, NUMERO, ...
    if len(parts) >= 2:
        result["company_name"] = parts[1].upper()

    if len(parts) >= 3:
        result["year"] = parts[2]

    if len(parts) >= 4:
        result["number"] = parts[3]

    if len(parts) >= 5:
        result["description"] = " ".join(parts[4:])

    return result

def normalize_sheet_token(sheet_name: str):
    return normalize_text(sheet_name).upper()


def detect_sheet_pattern(sheet_name: str):
    """
    Espera hojas tipo:
    IMPO TERR
    IMPO MAR
    IMPO AEREA
    EXPO TERR
    EXPO MAR
    EXPO AEREA
    """
    s = normalize_sheet_token(sheet_name)

    operation_type = None
    transport_key = None

    if "IMPO" in s:
        operation_type = "importacion"
    elif "EXPO" in s:
        operation_type = "exportacion"

    if "TERR" in s:
        transport_key = "terrestre"
    elif "AERE" in s:
        transport_key = "aereo"
    elif "MAR" in s:
        transport_key = "maritimo"

    if not operation_type or not transport_key:
        return None

    return {
        "operationType": operation_type,
        "transportKey": transport_key
    }


def operation_label(operation_type: str):
    return {
        "importacion": "Importaciones",
        "exportacion": "Exportaciones"
    }.get(operation_type, "Operaciones")


def transport_label(transport_key: str):
    return {
        "terrestre": "terrestres",
        "aereo": "aéreas",
        "maritimo": "marítimas"
    }.get(transport_key, "logísticas")


def build_friendly_sheet_label(company_name: str, operation_type: str, transport_key: str):
    return f"{(company_name or 'EMPRESA').upper()} · {operation_label(operation_type)} {transport_label(transport_key)}"

def normalize_header(value):
    text = normalize_text(value)
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.upper()


def is_blank_tc_value(value):
    if value is None:
        return True
    if isinstance(value, (int, float)):
        return float(value) == 0.0
    text = normalize_text(value)
    return text in {"", "-", "—", "0", "0.0", "0,0", "0,00"}


def parse_excel_like_date(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    # toma solo la parte fecha si viene con hora
    match = re.search(r"(\d{1,2}[/-]\d{1,2}[/-]\d{4})", text)
    if match:
        date_part = match.group(1).replace("-", "/")
        try:
            return datetime.strptime(date_part, "%d/%m/%Y").date()
        except Exception:
            pass

    patterns = [
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    ]

    for fmt in patterns:
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass

    return None

def format_excel_display_value(value):
    if value is None:
        return ""

    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%d/%m/%Y")
        return value.strftime("%d/%m/%Y %H:%M")

    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    return normalize_text(value)

def format_clp_dollar(value_float):
    text = f"{value_float:,.2f}"
    text = text.replace(",", "X").replace(".", ",").replace("X", ".")
    return text

def canonicalize_value(value):
    if isinstance(value, dict):
        return {
            "text": normalize_text(value.get("text")),
            "url": normalize_text(value.get("url")),
            "tooltip": normalize_text(value.get("tooltip")),
        }
    if value is None:
        return ""
    return normalize_text(value)


def build_row_identity(file_name, sheet_name, values, headers, row_idx):
    ref_header = next(
        (h for h in headers if "REFERENCIA CLIENTE" in normalize_header(h)),
        headers[0] if headers else ""
    )
    dispatch_header = next(
        (h for h in headers if "NUMERO DE DESPACHO" in normalize_header(h) or normalize_header(h) == "DESPACHO"),
        headers[1] if len(headers) > 1 else ""
    )

    ref_val = canonicalize_value(values.get(ref_header, ""))
    dispatch_val = canonicalize_value(values.get(dispatch_header, ""))

    if isinstance(ref_val, dict):
        ref_val = ref_val.get("text", "")
    if isinstance(dispatch_val, dict):
        dispatch_val = dispatch_val.get("text", "")

    ref_val = normalize_text(ref_val)
    dispatch_val = normalize_text(dispatch_val)

    # Si faltan los identificadores principales, usamos fallback con fila Excel
    if not ref_val and not dispatch_val:
        return "|".join([
            normalize_text(file_name),
            normalize_text(sheet_name),
            f"ROW-{row_idx}"
        ])

    return "|".join([
        normalize_text(file_name),
        normalize_text(sheet_name),
        ref_val,
        dispatch_val,
    ])


def normalize_row_data_for_compare(row_data):
    normalized = {}
    for key, value in (row_data or {}).items():
        normalized[key] = canonicalize_value(value)
    return normalized


def diff_row_values(old_values, new_values):
    changes = []
    all_keys = sorted(set(old_values.keys()) | set(new_values.keys()))

    for key in all_keys:
        old_val = old_values.get(key)
        new_val = new_values.get(key)

        if old_val != new_val:
            changes.append({
                "column": key,
                "old": old_val,
                "new": new_val
            })

    return changes


def supabase_headers():
    return {
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }


def fetch_existing_snapshots(file_key, sheet_name):
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise RuntimeError("Faltan SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY")

    url = (
        f"{SUPABASE_URL}/rest/v1/row_snapshot"
        f"?file_key=eq.{requests.utils.quote(file_key)}"
        f"&sheet_name=eq.{requests.utils.quote(sheet_name)}"
    )
    res = requests.get(url, headers=supabase_headers(), timeout=60)
    res.raise_for_status()

    items = res.json() or []
    return {
        item["row_identity"]: item
        for item in items
    }


def upsert_snapshot(file_key, sheet_name, row_identity, excel_row_number, row_data):
    url = f"{SUPABASE_URL}/rest/v1/row_snapshot"
    payload = [{
        "file_key": file_key,
        "sheet_name": sheet_name,
        "row_identity": row_identity,
        "excel_row_number": excel_row_number,
        "row_data": row_data,
        "last_seen_at": datetime.utcnow().isoformat()
    }]
    headers = supabase_headers()
    headers["Prefer"] = "resolution=merge-duplicates"

    res = requests.post(url, headers=headers, data=json.dumps(payload), timeout=60)
    res.raise_for_status()


def delete_snapshot(file_key, sheet_name, row_identity):
    url = (
        f"{SUPABASE_URL}/rest/v1/row_snapshot"
        f"?file_key=eq.{requests.utils.quote(file_key)}"
        f"&sheet_name=eq.{requests.utils.quote(sheet_name)}"
        f"&row_identity=eq.{requests.utils.quote(row_identity)}"
    )
    res = requests.delete(url, headers=supabase_headers(), timeout=60)
    res.raise_for_status()


def insert_audit_log(file_key, sheet_name, row_identity, excel_row_number, event_type, summary, diff=None, changed_by="system", changed_by_label="Sistema"):
    url = f"{SUPABASE_URL}/rest/v1/row_audit_log"
    payload = [{
        "file_key": file_key,
        "sheet_name": sheet_name,
        "row_identity": row_identity,
        "excel_row_number": excel_row_number,
        "event_type": event_type,
        "changed_by": changed_by,
        "changed_by_label": changed_by_label,
        "summary": summary,
        "diff": diff or []
    }]
    res = requests.post(url, headers=supabase_headers(), data=json.dumps(payload), timeout=60)
    res.raise_for_status()


def parse_cmf_value(raw):
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    return float(text.replace(".", "").replace(",", "."))


def get_share_token_from_url(share_url: str):
    encoded = base64.b64encode(share_url.encode("utf-8")).decode("utf-8")
    encoded = encoded.rstrip("=").replace("/", "_").replace("+", "-")
    return f"u!{encoded}"


def download_shared_pdf_bytes(share_url: str, token: str):
    share_token = get_share_token_from_url(share_url)

    meta_url = f"https://graph.microsoft.com/v1.0/shares/{share_token}/driveItem"
    headers = {"Authorization": f"Bearer {token}"}
    meta_res = requests.get(meta_url, headers=headers, timeout=60)
    meta_res.raise_for_status()
    meta = meta_res.json()

    download_url = meta.get("@microsoft.graph.downloadUrl")
    if download_url:
        file_res = requests.get(download_url, timeout=120)
        file_res.raise_for_status()
        return file_res.content

    item_id = meta.get("id")
    parent_ref = meta.get("parentReference", {})
    drive_id = parent_ref.get("driveId")
    if not drive_id or not item_id:
        raise RuntimeError("No se pudo resolver el PDF compartido desde SharePoint.")

    content_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    file_res = requests.get(content_url, headers=headers, timeout=120)
    file_res.raise_for_status()
    return file_res.content


def extract_fecha_pago_from_pdf_bytes(pdf_bytes: bytes):
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "\n".join((page.extract_text() or "") for page in reader.pages)

    match = re.search(r"Fecha\s+Pago\s+(\d{2}-\d{2}-\d{4})(?:\s+\d{2}:\d{2}:\d{2})?", text, re.IGNORECASE)
    if not match:
        return None

    try:
        return datetime.strptime(match.group(1), "%d-%m-%Y").date()
    except Exception:
        return None


def fetch_cmf_dollar_for_date(target_date: date, cmf_cache: dict):
    if target_date is None:
        return None

    cache_key = target_date.isoformat()
    if cache_key in cmf_cache:
        return cmf_cache[cache_key]

    if not CMF_API_KEY:
        cmf_cache[cache_key] = None
        return None

    headers = {"User-Agent": "PirazzoliDemo/1.0"}

    # intento exacto
    exact_url = (
        f"https://api.cmfchile.cl/api-sbifv3/recursos_api/dolar/"
        f"{target_date.year}/{target_date.month:02d}/dias/{target_date.day:02d}"
        f"?apikey={CMF_API_KEY}&formato=json"
    )
    exact_res = requests.get(exact_url, headers=headers, timeout=60)

    if exact_res.ok:
        data = exact_res.json()
        dolares = data.get("Dolares", []) or data.get("Dolar", [])
        if dolares:
            value = parse_cmf_value(dolares[0].get("Valor"))
            if value is not None:
                formatted = format_clp_dollar(value)
                cmf_cache[cache_key] = formatted
                return formatted

    # fallback: fecha anterior disponible
    prev_url = (
        f"https://api.cmfchile.cl/api-sbifv3/recursos_api/dolar/anteriores/"
        f"{target_date.year}/{target_date.month:02d}/dias/{target_date.day:02d}"
        f"?apikey={CMF_API_KEY}&formato=json"
    )
    prev_res = requests.get(prev_url, headers=headers, timeout=60)
    prev_res.raise_for_status()
    data = prev_res.json()
    dolares = data.get("Dolares", []) or data.get("Dolar", [])

    best_date = None
    best_value = None

    for item in dolares:
        fecha_txt = item.get("Fecha")
        valor_txt = item.get("Valor")
        if not fecha_txt:
            continue
        try:
            item_date = datetime.strptime(fecha_txt, "%Y-%m-%d").date()
        except Exception:
            continue

        if item_date <= target_date:
            if best_date is None or item_date > best_date:
                best_date = item_date
                best_value = parse_cmf_value(valor_txt)

    if best_value is None:
        cmf_cache[cache_key] = None
        return None

    formatted = format_clp_dollar(best_value)
    cmf_cache[cache_key] = formatted
    return formatted

def get_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "client_credentials",
        "scope": "https://graph.microsoft.com/.default",
    }
    res = requests.post(url, data=data, timeout=30)
    res.raise_for_status()
    return res.json()["access_token"]


def graph_get(url, token):
    headers = {"Authorization": f"Bearer {token}"}
    res = requests.get(url, headers=headers, timeout=60)
    res.raise_for_status()
    return res.json()


def graph_get_bytes(url, token):
    headers = {"Authorization": f"Bearer {token}"}
    res = requests.get(url, headers=headers, timeout=120)
    res.raise_for_status()
    return res.content

EXCEL_EXTENSIONS = (".xlsx", ".xlsm", ".xls")

def is_excel_filename(name: str) -> bool:
    return str(name or "").lower().endswith(EXCEL_EXTENSIONS)

def fetch_excel_files(token=None):
    token = token or get_token()

    site = graph_get(
        f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_HOSTNAME}:{SHAREPOINT_SITE_PATH}",
        token
    )
    site_id = site["id"]

    drives = graph_get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives",
        token
    ).get("value", [])

    available_drive_names = [d.get("name") for d in drives]

    target_drive = next((d for d in drives if d.get("name") == TARGET_DRIVE_NAME), None)
    if not target_drive:
        raise RuntimeError(
            f"No se encontró el drive '{TARGET_DRIVE_NAME}'. Disponibles: {available_drive_names}"
        )

    drive_id = target_drive["id"]

    children = graph_get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/children",
        token
    ).get("value", [])

    excel_items = [
        i for i in children
        if is_excel_filename(i.get("name"))
    ]

    if not excel_items:
        available_root_items = [i.get("name") for i in children]
        raise RuntimeError(
            f"No se encontraron archivos Excel en la raíz del drive. Elementos en raíz: {available_root_items}"
        )

    results = []
    for item in excel_items:
        file_id = item["id"]
        file_name = item["name"]
        last_modified = item.get("lastModifiedDateTime")

        content = graph_get_bytes(
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/content",
            token
        )

        results.append({
            "file_name": file_name,
            "last_modified": last_modified,
            "content": content
        })

    return results


def fetch_excel_bytes(token=None):
    token = token or get_token()

    site = graph_get(
        f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_HOSTNAME}:{SHAREPOINT_SITE_PATH}",
        token
    )
    site_id = site["id"]

    drives = graph_get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives",
        token
    ).get("value", [])

    available_drive_names = [d.get("name") for d in drives]

    target_drive = next((d for d in drives if d.get("name") == TARGET_DRIVE_NAME), None)
    if not target_drive:
        raise RuntimeError(
            f"No se encontró el drive '{TARGET_DRIVE_NAME}'. Disponibles: {available_drive_names}"
        )

    drive_id = target_drive["id"]

    children = graph_get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/children",
        token
    ).get("value", [])

    available_root_items = [i.get("name") for i in children]

    target_file = next((i for i in children if i.get("name") == TARGET_FILE_NAME), None)
    if not target_file:
        raise RuntimeError(
            f"No se encontró el archivo '{TARGET_FILE_NAME}' en la raíz del drive. Elementos en raíz: {available_root_items}"
        )

    file_id = target_file["id"]
    last_modified = target_file.get("lastModifiedDateTime")

    content = graph_get_bytes(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{file_id}/content",
        token
    )

    return content, last_modified

def enrich_exchange_rate_columns(rows, headers, token):
    normalized_map = {normalize_header(h): h for h in headers}

    hdr_pago_tgr = next((h for n, h in normalized_map.items() if "COMPROBANTE PAGO TGR" in n), None)
    hdr_fecha_liberacion = next((h for n, h in normalized_map.items() if "FECHA DE LIBERACION" in n), None)
    hdr_tc_legalizacion = next((h for n, h in normalized_map.items() if "T/C DOLAR LEGALIZACION" in n), None)
    hdr_tc_liberacion = next((h for n, h in normalized_map.items() if "T/C LIBERACION CARGA" in n), None)

    if not hdr_tc_legalizacion and not hdr_tc_liberacion:
        return rows

    pdf_date_cache = PDF_DATE_CACHE
    cmf_cache = CMF_RATE_CACHE

    for row in rows:
        values = row["values"]

        # AK - T/C DOLAR legalizacion
        if hdr_tc_legalizacion and is_blank_tc_value(values.get(hdr_tc_legalizacion)):
            source_date = None

            pago_val = values.get(hdr_pago_tgr) if hdr_pago_tgr else None

            # 1) primero intenta usar la fecha visible en AF
            if isinstance(pago_val, dict):
                source_date = parse_excel_like_date(pago_val.get("text"))
            else:
                source_date = parse_excel_like_date(pago_val)

            # 2) si no hay fecha visible, intenta leer el PDF
            if source_date is None and isinstance(pago_val, dict) and pago_val.get("url"):
                pdf_url = pago_val["url"]

                if pdf_url in pdf_date_cache:
                    source_date = pdf_date_cache[pdf_url]
                else:
                    try:
                        pdf_bytes = download_shared_pdf_bytes(pdf_url, token)
                        source_date = extract_fecha_pago_from_pdf_bytes(pdf_bytes)
                    except Exception:
                        source_date = None

                    pdf_date_cache[pdf_url] = source_date

            if source_date is not None:
                dolar_val = fetch_cmf_dollar_for_date(source_date, cmf_cache)
                if dolar_val is not None:
                    values[hdr_tc_legalizacion] = dolar_val

        # AL - T/C liberacion carga
        if hdr_tc_liberacion and is_blank_tc_value(values.get(hdr_tc_liberacion)):
            liberacion_val = values.get(hdr_fecha_liberacion) if hdr_fecha_liberacion else None
            source_date = parse_excel_like_date(liberacion_val)

            if source_date is not None:
                dolar_val = fetch_cmf_dollar_for_date(source_date, cmf_cache)
                if dolar_val is not None:
                    values[hdr_tc_liberacion] = dolar_val

    return rows

def parse_workbook_to_payload(content: bytes, token: str, file_name: str, run_audit: bool = False):
    wb_values = load_workbook(io.BytesIO(content), data_only=True)
    wb_raw = load_workbook(io.BytesIO(content), data_only=False)

    file_meta = parse_excel_file_name(file_name)
    company_name = file_meta["company_name"]

    files = [{
        "file": file_name,
        "baseName": file_meta["base_name"],
        "companyName": company_name,
        "fileMeta": {
            "prefix": file_meta["prefix"],
            "year": file_meta["year"],
            "number": file_meta["number"],
            "description": file_meta["description"],
        },
        "sheets": []
    }]

    for sheet_name in wb_values.sheetnames:
        ws_values = wb_values[sheet_name]
        ws_raw = wb_raw[sheet_name]

        if sheet_name.strip().upper() in {"MATRIZ DATOS", "HOJA1"}:
            continue

        sheet_meta = detect_sheet_pattern(sheet_name)
        if not sheet_meta:
            continue

        transport_key = sheet_meta["transportKey"]
        operation_type = sheet_meta["operationType"]
        friendly_name = build_friendly_sheet_label(company_name, operation_type, transport_key)

        headers = []

        headers = []
        header_meta = []
        visible_columns = []

        for col_idx, cell in enumerate(ws_values[1], start=1):
            header = normalize_text(cell.value)
            if not header:
                header = f"Columna {col_idx}"
            headers.append(header)
            header_meta.append({
                "header": header,
                "colIndex": col_idx - 1,
                "excelColLetter": cell.column_letter
            })
            visible_columns.append((col_idx, header, cell.column_letter))

        rows = []
        for row_idx in range(2, ws_values.max_row + 1):
            values = {}
            has_any = False

            for col_idx, header, _letter in visible_columns:
                cell_value = ws_values.cell(row=row_idx, column=col_idx)
                cell_raw = ws_raw.cell(row=row_idx, column=col_idx)

                raw_display = cell_value.value
                text = format_excel_display_value(raw_display)

                hyperlink = None
                if cell_raw.hyperlink and cell_raw.hyperlink.target:
                    hyperlink = cell_raw.hyperlink.target

                if hyperlink:
                    values[header] = {
                        "text": text or header,
                        "url": hyperlink,
                        "tooltip": ""
                    }
                    has_any = True
                else:
                    values[header] = text
                    if text not in ("", "-", "—"):
                        has_any = True

            if not has_any:
                continue

            ref_header = next((h for h in headers if "REFERENCIA CLIENTE" in h.upper()), headers[0] if headers else "")
            dispatch_header = next((h for h in headers if "NUMERO DE DESPACHO" in h.upper() or h.upper() == "DESPACHO"), headers[1] if len(headers) > 1 else "")

            base_id = "|".join([
                file_name,
                sheet_name,
                str(values.get(ref_header, "—")),
                str(values.get(dispatch_header, "—")),
                str(row_idx)
            ])

            row_identity = build_row_identity(file_name, sheet_name, values, headers, row_idx)

            rows.append({
                "_id": base_id,
                "_rowIdentity": row_identity,
                "_sheet": sheet_name,
                "_file": file_name,
                "_company": company_name,
                "_excelRowNumber": row_idx,
                "values": values
            })

        rows = enrich_exchange_rate_columns(rows, headers, token)

        if run_audit:
            try:
                previous_snapshots = fetch_existing_snapshots(file_name, sheet_name)
            except Exception as e:
                previous_snapshots = {}
                print(f"[AUDIT] No se pudieron cargar snapshots previos de {sheet_name}: {e}")

            current_identities = set()

            for row in rows:
                row_identity = row["_rowIdentity"]
                current_identities.add(row_identity)

                current_row_number = row["_excelRowNumber"]
                current_row_data = normalize_row_data_for_compare(row["values"])

                previous = previous_snapshots.get(row_identity)

                if not previous:
                    try:
                        insert_audit_log(
                            file_key=file_name,
                            sheet_name=sheet_name,
                            row_identity=row_identity,
                            excel_row_number=current_row_number,
                            event_type="created",
                            summary="Fila creada desde el Excel",
                            diff=[]
                        )
                    except Exception as e:
                        print(f"[AUDIT] Error guardando created: {e}")

                    try:
                        upsert_snapshot(
                            file_key=file_name,
                            sheet_name=sheet_name,
                            row_identity=row_identity,
                            excel_row_number=current_row_number,
                            row_data=current_row_data
                        )
                    except Exception as e:
                        print(f"[AUDIT] Error guardando snapshot nuevo: {e}")

                    continue

                old_row_number = previous.get("excel_row_number")
                old_row_data = normalize_row_data_for_compare(previous.get("row_data") or {})
                changes = diff_row_values(old_row_data, current_row_data)

                if changes:
                    try:
                        insert_audit_log(
                            file_key=file_name,
                            sheet_name=sheet_name,
                            row_identity=row_identity,
                            excel_row_number=current_row_number,
                            event_type="updated",
                            summary=f"Se modificaron {len(changes)} campo(s)",
                            diff=changes
                        )
                    except Exception as e:
                        print(f"[AUDIT] Error guardando updated: {e}")

                elif old_row_number != current_row_number:
                    try:
                        insert_audit_log(
                            file_key=file_name,
                            sheet_name=sheet_name,
                            row_identity=row_identity,
                            excel_row_number=current_row_number,
                            event_type="moved",
                            summary=f"La fila cambió de posición: {old_row_number} → {current_row_number}",
                            diff=[]
                        )
                    except Exception as e:
                        print(f"[AUDIT] Error guardando moved: {e}")

                try:
                    upsert_snapshot(
                        file_key=file_name,
                        sheet_name=sheet_name,
                        row_identity=row_identity,
                        excel_row_number=current_row_number,
                        row_data=current_row_data
                    )
                except Exception as e:
                    print(f"[AUDIT] Error actualizando snapshot: {e}")

            deleted_identities = set(previous_snapshots.keys()) - current_identities

            for deleted_identity in deleted_identities:
                previous = previous_snapshots[deleted_identity]
                try:
                    insert_audit_log(
                        file_key=file_name,
                        sheet_name=sheet_name,
                        row_identity=deleted_identity,
                        excel_row_number=previous.get("excel_row_number"),
                        event_type="deleted",
                        summary="Fila eliminada del Excel",
                        diff=[]
                    )
                except Exception as e:
                    print(f"[AUDIT] Error guardando deleted: {e}")

                try:
                    delete_snapshot(
                        file_key=file_name,
                        sheet_name=sheet_name,
                        row_identity=deleted_identity
                    )
                except Exception as e:
                    print(f"[AUDIT] Error eliminando snapshot: {e}")

        files[0]["sheets"].append({
            "name": sheet_name,
            "headers": headers,
            "headerMeta": header_meta,
            "rows": rows,
            "transportKey": transport_key,
            "operationType": operation_type,
            "companyName": company_name,
            "friendlyName": friendly_name
        })

    return files


@app.route("/api", methods=["GET"])
def api_root():
    try:
        force_refresh = str(request.args.get("refresh", "")).lower() in {"1", "true", "yes"}
        return jsonify(build_payload(force_refresh=force_refresh))
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500


@app.route("/api/row-history", methods=["GET"])
def api_row_history():
    try:
        file_key = request.args.get("file_key", "").strip()
        sheet_name = request.args.get("sheet_name", "").strip()
        row_identity = request.args.get("row_identity", "").strip()

        if not file_key or not sheet_name or not row_identity:
            return jsonify({
                "ok": False,
                "error": "Faltan parámetros obligatorios."
            }), 400

        url = (
            f"{SUPABASE_URL}/rest/v1/row_audit_log"
            f"?file_key=eq.{requests.utils.quote(file_key)}"
            f"&sheet_name=eq.{requests.utils.quote(sheet_name)}"
            f"&row_identity=eq.{requests.utils.quote(row_identity)}"
            f"&order=changed_at.desc"
        )

        res = requests.get(url, headers=supabase_headers(), timeout=60)
        res.raise_for_status()

        return jsonify({
            "ok": True,
            "items": res.json() or []
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500