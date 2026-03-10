# -*- coding: utf-8 -*-
import os
import json
import math
import openpyxl
from datetime import datetime
from urllib.parse import urlparse, parse_qs, unquote

BASE = os.path.dirname(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE, "data")
OUT_PATH = os.path.join(DATA_DIR, "despachos.json")
EXACT_EXCEL = os.path.join(BASE, "piloto_datos_minerva2.xlsx")

def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    return str(value).strip()

def normalize(value):
    return clean_text(value).strip().upper()

def format_date_like_excel(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return clean_text(value)

def find_excel(base):
    if os.path.exists(EXACT_EXCEL):
        return EXACT_EXCEL

    candidatos = []
    for f in os.listdir(base):
        if f.lower().endswith(".xlsx") and not f.startswith("~$"):
            candidatos.append(os.path.join(base, f))
    if not candidatos:
        raise Exception("No se encontró un archivo .xlsx en la raíz del proyecto.")
    candidatos.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return candidatos[0]

def get_hyperlink_or_value(cell, base_dir):
    text = clean_text(cell.value)
    url = None

    if cell.hyperlink and cell.hyperlink.target:
        url = cell.hyperlink.target
    elif isinstance(cell.value, str) and cell.value.strip().lower().startswith("http"):
        url = cell.value.strip()

    local_url = None
    filename = None

    if url:
        try:
            parsed = urlparse(url)
            qs = parse_qs(parsed.query)

            if "id" in qs and qs["id"]:
                path_value = unquote(qs["id"][0])
                filename = os.path.basename(path_value)
            else:
                filename = os.path.basename(parsed.path)

            if filename:
                local_path = os.path.join(base_dir, filename)
                if os.path.exists(local_path):
                    local_url = f"./{filename}"
        except Exception:
            pass

    return {
        "text": text,
        "url": url,
        "localUrl": local_url,
        "filename": filename
    }

def estado_doc_from_cell(value, has_link=False):
    v = normalize(value)

    if has_link and v == "":
        return "Presentado"
    if "APROB" in v or "LIBER" in v or v == "OK":
        return "Aprobado"
    if "PEND" in v:
        return "Pendiente"
    if "OBS" in v:
        return "Observado"
    if "PRES" in v:
        return "Presentado"
    if has_link:
        return "Presentado"
    return clean_text(value)

def value_or_empty(ws, row, headers, key):
    if key not in headers:
        return ""
    return ws.cell(row, headers[key]).value

def doc_obj(info, label):
    if not info["url"] and not info["text"]:
        return None
    return {
        "label": label,
        "text": info["text"] if info["text"] else f"Abrir {label}",
        "url": info["url"],
        "localUrl": info["localUrl"],
        "filename": info["filename"]
    }

def find_header_exact_or_trim(headers, target):
    for key in headers.keys():
        if key.strip().upper() == target.strip().upper():
            return key
    return None

excel_path = find_excel(BASE)
print("Excel encontrado:", excel_path)

wb = openpyxl.load_workbook(excel_path, data_only=False)

if "OPERACIONES" not in wb.sheetnames:
    raise Exception("La hoja 'OPERACIONES' no existe en el Excel.")

ws = wb["OPERACIONES"]

HEADER_ROW = 1
DATA_START_ROW = 2

headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(HEADER_ROW, col).value
    if val is not None:
        headers[str(val).strip()] = col

print("Encabezados encontrados:", len(headers))
print("Máx filas:", ws.max_row)
print("Máx columnas:", ws.max_column)

comp_pago_key = find_header_exact_or_trim(headers, "COMPROBANTE PAGO TGR")

rows = []

for row in range(DATA_START_ROW, ws.max_row + 1):
    numero_despacho = clean_text(value_or_empty(ws, row, headers, "Número de despacho."))
    referencia_cliente = clean_text(value_or_empty(ws, row, headers, "Referencia de cliente."))

    if numero_despacho == "" and referencia_cliente == "":
        continue

    estado_carga = clean_text(value_or_empty(ws, row, headers, "ESTADO DE LA CARGA"))

    info_cda = get_hyperlink_or_value(ws.cell(row, headers["CDA SEREMI"]), BASE) if "CDA SEREMI" in headers else {"text":"","url":None,"localUrl":None,"filename":None}
    info_res = get_hyperlink_or_value(ws.cell(row, headers["Resolucion Sanitaria UYD"]), BASE) if "Resolucion Sanitaria UYD" in headers else {"text":"","url":None,"localUrl":None,"filename":None}
    info_sag = get_hyperlink_or_value(ws.cell(row, headers["SAG"]), BASE) if "SAG" in headers else {"text":"","url":None,"localUrl":None,"filename":None}
    info_iipa = get_hyperlink_or_value(ws.cell(row, headers["IIPA"]), BASE) if "IIPA" in headers else {"text":"","url":None,"localUrl":None,"filename":None}
    info_din = get_hyperlink_or_value(ws.cell(row, headers["DIN"]), BASE) if "DIN" in headers else {"text":"","url":None,"localUrl":None,"filename":None}
    info_tgr = get_hyperlink_or_value(ws.cell(row, headers[comp_pago_key]), BASE) if comp_pago_key else {"text":"","url":None,"localUrl":None,"filename":None}

    cda_estado = estado_doc_from_cell(value_or_empty(ws, row, headers, "CDA SEREMI"), has_link=bool(info_cda["url"]))
    res_estado = estado_doc_from_cell(value_or_empty(ws, row, headers, "Resolucion Sanitaria UYD"), has_link=bool(info_res["url"]))
    sag_estado = estado_doc_from_cell(value_or_empty(ws, row, headers, "SAG"), has_link=bool(info_sag["url"]))
    iipa_estado = estado_doc_from_cell(value_or_empty(ws, row, headers, "IIPA"), has_link=bool(info_iipa["url"]))
    din_estado = estado_doc_from_cell(value_or_empty(ws, row, headers, "DIN"), has_link=bool(info_din["url"]))

    item = {
        "_id": f"row-{row}",

        "numeroDespacho": numero_despacho,
        "referenciaCliente": referencia_cliente,
        "estadoCarga": estado_carga,
        "observaciones": clean_text(value_or_empty(ws, row, headers, "Observaciones")),

        "confirmacionRecepcionDoc": format_date_like_excel(value_or_empty(ws, row, headers, "Confirmación de recepción de documentación (fecha correo)")),
        "estadoDocumentos": clean_text(value_or_empty(ws, row, headers, "Indicación de si los documentos se encuentran OK o con observaciones.")),
        "fechaSolicitudModificacion": format_date_like_excel(value_or_empty(ws, row, headers, "Fecha en que se solicitó modificación.")),
        "fechaNotificacion": format_date_like_excel(value_or_empty(ws, row, headers, "Fecha en que se notificó.")),
        "fechaCorrecciones": format_date_like_excel(value_or_empty(ws, row, headers, "Fecha en que se recibieron correcciones.")),
        "fechaNuevaPresentacion": format_date_like_excel(value_or_empty(ws, row, headers, "Fecha de nueva presentación, si corresponde.")),

        "tipoCarga": clean_text(value_or_empty(ws, row, headers, "TIPO CARGA")),
        "origenCarga": clean_text(value_or_empty(ws, row, headers, "Origen de la carga.")),
        "tipoOperacion": clean_text(value_or_empty(ws, row, headers, "Tipo operacion")),
        "pasoAduanero": clean_text(value_or_empty(ws, row, headers, "Paso aduanero (cuando corresponda).")),
        "facturaComercial": clean_text(value_or_empty(ws, row, headers, "factura comercial")),
        "micCrt": clean_text(value_or_empty(ws, row, headers, "MIC / CRT.")),
        "patenteTracto": clean_text(value_or_empty(ws, row, headers, "Patente del camión (TRACTO)")),
        "patenteRemolque": clean_text(value_or_empty(ws, row, headers, "Patente del camión (REMOLQUE)")),
        "numeroContenedor": clean_text(value_or_empty(ws, row, headers, "N° CONTENEDOR ")),
        "clienteFinal": clean_text(value_or_empty(ws, row, headers, "CLIENTE FINAL")),
        "direccionDescarga": clean_text(value_or_empty(ws, row, headers, "Dirección de descarga que saldrá en la guía. DESTINO")),
        "chofer": clean_text(value_or_empty(ws, row, headers, "Nombre del chofer e identificación con la que ingresó.")),
        "estadoInstruccionDescarga": clean_text(value_or_empty(ws, row, headers, "Estado: INSTRUCCIÓN de DESCARGA")),

        "cdaSeremiEstado": cda_estado,
        "resolucionSanitariaEstado": res_estado,
        "sagEstado": sag_estado,
        "iipaEstado": iipa_estado,
        "fechaAceptacionDin": format_date_like_excel(value_or_empty(ws, row, headers, "Fecha de aceptacion DIN")),
        "dinEstado": din_estado,

        "fechaIngresoAlmacenista": format_date_like_excel(value_or_empty(ws, row, headers, "Fecha de ingreso ALMACENISTA")),
        "fechaPapeletaIngreso": format_date_like_excel(value_or_empty(ws, row, headers, "FECHA PAPELETA INGRESO")),
        "fechaPresentacionDespacho": format_date_like_excel(value_or_empty(ws, row, headers, "Fecha de presentación de documentos.( DESPACHO)")),
        "fechaLiberacion": format_date_like_excel(value_or_empty(ws, row, headers, "Fecha de liberación.( SALIDA DE ALMACEN)")),
        "selloAga": clean_text(value_or_empty(ws, row, headers, "SELLO AGA")),
        "selloPuerto": clean_text(value_or_empty(ws, row, headers, "SELLO PUERTO")),
        "facturaAgencia": clean_text(value_or_empty(ws, row, headers, "FACTURA AGENCIA ")),
        "fechaFacturaAgencia": format_date_like_excel(value_or_empty(ws, row, headers, "FECHA FACTURA AGENCIA")),
        "remesa": clean_text(value_or_empty(ws, row, headers, "REMESA")),
        "pagoFactura": clean_text(value_or_empty(ws, row, headers, "PAGO FACTURA")),
        "tcdolarLegalizacion": clean_text(value_or_empty(ws, row, headers, "T/C DOLAR legalizacion")),
        "tcLiberacionCarga": clean_text(value_or_empty(ws, row, headers, "T/C liberacion carga")),
        "estadoOperacion": clean_text(value_or_empty(ws, row, headers, "ESTADO OPERACION ")),

        "docCdaSeremi": doc_obj(info_cda, "CDA / SEREMI"),
        "docResolucionSanitaria": doc_obj(info_res, "Resolución Sanitaria"),
        "docSag": doc_obj(info_sag, "SAG"),
        "docIipa": doc_obj(info_iipa, "IIPA"),
        "docDin": doc_obj(info_din, "DIN"),
        "docComprobanteTgr": doc_obj(info_tgr, "Comprobante pago TGR")
    }

    rows.append(item)

os.makedirs(DATA_DIR, exist_ok=True)
with open(OUT_PATH, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print("JSON generado:", OUT_PATH)
print("Filas exportadas:", len(rows))

# Mostrar conteo de estados para verificar
conteo = {}
for r in rows:
    est = normalize(r.get("estadoCarga", ""))
    conteo[est] = conteo.get(est, 0) + 1

print("Conteo por estado:")
for k, v in sorted(conteo.items()):
    print(" -", k, ":", v)
